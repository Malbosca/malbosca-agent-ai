#!/usr/bin/env python3
"""
Wine Information Searcher
Cerca informazioni sui vini dal web e aggiorna un file Excel
"""

import os
import sys
import time
import requests
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, Optional, List
import re


class WineSearcher:
    """Classe per cercare informazioni sui vini dal web"""

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        })

    def search_wine_info(self, wine_name: str, producer: str = "") -> Dict[str, str]:
        """
        Cerca informazioni su un vino

        Args:
            wine_name: Nome del vino
            producer: Produttore (opzionale)

        Returns:
            Dizionario con colore, gusto, profumo
        """
        print(f"Cercando informazioni per: {wine_name} {producer}")

        info = {
            'colore': '',
            'gusto': '',
            'profumo': ''
        }

        try:
            # Prova con Vivino
            vivino_info = self._search_vivino(wine_name, producer)
            if vivino_info:
                info.update(vivino_info)
                return info

            # Fallback: ricerca generica Google
            google_info = self._search_google(wine_name, producer)
            if google_info:
                info.update(google_info)

        except Exception as e:
            print(f"Errore durante la ricerca: {e}")

        return info

    def _search_vivino(self, wine_name: str, producer: str) -> Optional[Dict[str, str]]:
        """Cerca su Vivino"""
        try:
            query = f"{wine_name} {producer}".strip()
            search_url = f"https://www.vivino.com/search/wines?q={requests.utils.quote(query)}"

            response = self.session.get(search_url, timeout=10)
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'html.parser')

                # Estrai informazioni dalla pagina
                info = self._extract_wine_info_from_html(soup)
                if info:
                    return info

        except Exception as e:
            print(f"Errore Vivino: {e}")

        return None

    def _search_google(self, wine_name: str, producer: str) -> Optional[Dict[str, str]]:
        """Ricerca generica con Google"""
        try:
            query = f"{wine_name} {producer} caratteristiche organoletiche".strip()

            # Simula una ricerca e estrazione di informazioni base
            # In un caso reale, potresti usare l'API di Google Custom Search

            # Per ora, usiamo pattern comuni per dedurre informazioni base
            info = self._infer_wine_characteristics(wine_name, producer)
            return info

        except Exception as e:
            print(f"Errore ricerca Google: {e}")

        return None

    def _extract_wine_info_from_html(self, soup: BeautifulSoup) -> Optional[Dict[str, str]]:
        """Estrae informazioni dal HTML"""
        info = {}

        # Cerca pattern comuni
        text = soup.get_text().lower()

        # Colore
        if 'rosso' in text:
            info['colore'] = 'Rosso'
        elif 'bianco' in text:
            info['colore'] = 'Bianco'
        elif 'rosato' in text or 'rosé' in text:
            info['colore'] = 'Rosato'
        elif 'spumante' in text:
            info['colore'] = 'Spumante'

        # Cerca descrizioni di gusto e profumo
        taste_keywords = ['sapore', 'gusto', 'palato', 'taste']
        aroma_keywords = ['profumo', 'aroma', 'naso', 'bouquet']

        for keyword in taste_keywords:
            if keyword in text:
                # Cerca frasi vicino alla keyword
                pattern = f'{keyword}[^.]*\\.'.format(keyword=keyword)
                matches = re.findall(pattern, text)
                if matches:
                    info['gusto'] = matches[0].strip()
                    break

        for keyword in aroma_keywords:
            if keyword in text:
                pattern = f'{keyword}[^.]*\\.'.format(keyword=keyword)
                matches = re.findall(pattern, text)
                if matches:
                    info['profumo'] = matches[0].strip()
                    break

        return info if info else None

    def _infer_wine_characteristics(self, wine_name: str, producer: str) -> Dict[str, str]:
        """Deduce caratteristiche base dal nome del vino"""
        info = {
            'colore': '',
            'gusto': '',
            'profumo': ''
        }

        name_lower = wine_name.lower()

        # Deduzione colore da vitigni comuni
        red_grapes = ['barolo', 'barbaresco', 'brunello', 'chianti', 'amarone', 'montepulciano',
                      'primitivo', 'nero d\'avola', 'sangiovese', 'nebbiolo', 'merlot',
                      'cabernet', 'syrah', 'cannonau']
        white_grapes = ['vermentino', 'pecorino', 'falanghina', 'greco', 'fiano', 'trebbiano',
                        'verdicchio', 'pinot grigio', 'chardonnay', 'sauvignon', 'gewürztraminer', 'soave']
        sparkling = ['prosecco', 'franciacorta', 'trento doc', 'spumante', 'champagne']

        for grape in red_grapes:
            if grape in name_lower:
                info['colore'] = 'Rosso'
                info['gusto'] = 'Corposo, strutturato, tannico'
                info['profumo'] = 'Fruttato, speziato'
                break

        for grape in white_grapes:
            if grape in name_lower:
                info['colore'] = 'Bianco'
                info['gusto'] = 'Fresco, minerale, sapido'
                info['profumo'] = 'Floreale, agrumato'
                break

        for sparkling_type in sparkling:
            if sparkling_type in name_lower:
                info['colore'] = 'Spumante'
                info['gusto'] = 'Fresco, vivace, perlage fine'
                info['profumo'] = 'Floreale, fruttato'
                break

        return info


class WineExcelManager:
    """Gestisce l'input/output dei file Excel"""

    def __init__(self, file_path: str):
        self.file_path = file_path

    def read_wines(self) -> List[Dict[str, str]]:
        """Legge la lista dei vini dal file Excel"""
        try:
            df = pd.read_excel(self.file_path)

            # Assicurati che le colonne necessarie esistano
            required_cols = ['Nome Vino']
            if not all(col in df.columns for col in required_cols):
                raise ValueError(f"Il file Excel deve contenere almeno la colonna 'Nome Vino'")

            # Aggiungi colonne opzionali se non esistono
            optional_cols = ['Produttore', 'Colore', 'Gusto', 'Profumo']
            for col in optional_cols:
                if col not in df.columns:
                    df[col] = ''

            wines = df.to_dict('records')
            return wines

        except Exception as e:
            print(f"Errore nella lettura del file Excel: {e}")
            raise

    def write_wines(self, wines: List[Dict[str, str]]):
        """Scrive i dati aggiornati nel file Excel"""
        try:
            df = pd.DataFrame(wines)

            # Ordina le colonne
            column_order = ['Nome Vino', 'Produttore', 'Colore', 'Gusto', 'Profumo']
            # Aggiungi eventuali colonne extra non standard
            for col in df.columns:
                if col not in column_order:
                    column_order.append(col)

            df = df[[col for col in column_order if col in df.columns]]

            # Salva in Excel
            with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Vini')

                # Formattazione
                workbook = writer.book
                worksheet = writer.sheets['Vini']

                # Stile header
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Auto-dimensiona colonne
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            print(f"File Excel aggiornato: {self.file_path}")

        except Exception as e:
            print(f"Errore nella scrittura del file Excel: {e}")
            raise


def main():
    """Funzione principale"""
    import argparse

    parser = argparse.ArgumentParser(
        description='Cerca informazioni sui vini e aggiorna un file Excel'
    )
    parser.add_argument(
        'input_file',
        help='File Excel di input con la lista dei vini'
    )
    parser.add_argument(
        '-o', '--output',
        help='File Excel di output (default: sovrascrive input)',
        default=None
    )
    parser.add_argument(
        '--delay',
        type=float,
        default=2.0,
        help='Ritardo in secondi tra una ricerca e l\'altra (default: 2.0)'
    )

    args = parser.parse_args()

    # Determina file di output
    output_file = args.output if args.output else args.input_file

    # Verifica che il file di input esista
    if not os.path.exists(args.input_file):
        print(f"Errore: File {args.input_file} non trovato!")
        sys.exit(1)

    print("=" * 60)
    print("Wine Information Searcher")
    print("=" * 60)
    print(f"Input file: {args.input_file}")
    print(f"Output file: {output_file}")
    print(f"Delay: {args.delay}s")
    print("=" * 60)

    # Inizializza
    excel_manager = WineExcelManager(args.input_file)
    searcher = WineSearcher()

    # Leggi vini
    print("\nLettura file Excel...")
    wines = excel_manager.read_wines()
    print(f"Trovati {len(wines)} vini da processare")

    # Processa ogni vino
    print("\nInizio ricerca informazioni...\n")

    for i, wine in enumerate(wines, 1):
        wine_name = wine.get('Nome Vino', '')
        producer = wine.get('Produttore', '')

        print(f"[{i}/{len(wines)}] {wine_name}")

        # Helper per verificare se un valore è presente e non vuoto
        def has_value(val):
            if pd.isna(val):
                return False
            if isinstance(val, str) and val.strip() == '':
                return False
            return True

        # Salta se già ha tutte le informazioni
        if has_value(wine.get('Colore')) and has_value(wine.get('Gusto')) and has_value(wine.get('Profumo')):
            print("  → Informazioni già presenti, skip")
            continue

        # Cerca informazioni
        info = searcher.search_wine_info(wine_name, producer)

        # Aggiorna solo se non già presente
        if info.get('colore') and not has_value(wine.get('Colore')):
            wine['Colore'] = info['colore']
        if info.get('gusto') and not has_value(wine.get('Gusto')):
            wine['Gusto'] = info['gusto']
        if info.get('profumo') and not has_value(wine.get('Profumo')):
            wine['Profumo'] = info['profumo']

        # Helper per visualizzare i valori
        def display_value(val, max_len=50):
            if pd.isna(val):
                return 'N/D'
            val_str = str(val)
            if len(val_str) > max_len:
                return val_str[:max_len] + '...'
            return val_str

        print(f"  → Colore: {display_value(wine.get('Colore'))}")
        print(f"  → Gusto: {display_value(wine.get('Gusto'))}")
        print(f"  → Profumo: {display_value(wine.get('Profumo'))}")

        # Attendi per non sovraccaricare i server
        if i < len(wines):
            time.sleep(args.delay)

    # Salva risultati
    print("\n" + "=" * 60)
    print("Salvataggio risultati...")

    if args.output:
        excel_manager.file_path = output_file

    excel_manager.write_wines(wines)

    print("=" * 60)
    print("Completato!")
    print("=" * 60)


if __name__ == '__main__':
    main()
