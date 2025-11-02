#!/usr/bin/env python3
"""
Script per creare un file Excel di esempio con una lista di vini
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_example_excel(output_file='vini_esempio.xlsx'):
    """Crea un file Excel di esempio con vini italiani"""

    # Lista di vini di esempio
    wines = [
        {'Nome Vino': 'Barolo DOCG', 'Produttore': 'Marchesi di Barolo', 'Colore': '', 'Gusto': '', 'Profumo': ''},
        {'Nome Vino': 'Brunello di Montalcino DOCG', 'Produttore': 'Banfi', 'Colore': '', 'Gusto': '', 'Profumo': ''},
        {'Nome Vino': 'Amarone della Valpolicella', 'Produttore': 'Masi', 'Colore': '', 'Gusto': '', 'Profumo': ''},
        {'Nome Vino': 'Chianti Classico DOCG', 'Produttore': 'Antinori', 'Colore': '', 'Gusto': '', 'Profumo': ''},
        {'Nome Vino': 'Prosecco Superiore DOCG', 'Produttore': 'Bisol', 'Colore': '', 'Gusto': '', 'Profumo': ''},
        {'Nome Vino': 'Vermentino di Sardegna', 'Produttore': 'Sella & Mosca', 'Colore': '', 'Gusto': '', 'Profumo': ''},
        {'Nome Vino': 'Nero d\'Avola DOC', 'Produttore': 'Planeta', 'Colore': '', 'Gusto': '', 'Profumo': ''},
        {'Nome Vino': 'Franciacorta Brut', 'Produttore': 'Ca\' del Bosco', 'Colore': '', 'Gusto': '', 'Profumo': ''},
        {'Nome Vino': 'Primitivo di Manduria', 'Produttore': 'Feudi di San Marzano', 'Colore': '', 'Gusto': '', 'Profumo': ''},
        {'Nome Vino': 'Soave Classico DOC', 'Produttore': 'Pieropan', 'Colore': '', 'Gusto': '', 'Profumo': ''},
    ]

    # Crea DataFrame
    df = pd.DataFrame(wines)

    # Salva in Excel con formattazione
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Vini')

        # Formattazione
        workbook = writer.book
        worksheet = writer.sheets['Vini']

        # Stile header
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-dimensiona colonne
        worksheet.column_dimensions['A'].width = 30  # Nome Vino
        worksheet.column_dimensions['B'].width = 25  # Produttore
        worksheet.column_dimensions['C'].width = 15  # Colore
        worksheet.column_dimensions['D'].width = 40  # Gusto
        worksheet.column_dimensions['E'].width = 40  # Profumo

        # Allineamento celle
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    print(f"File Excel di esempio creato: {output_file}")

if __name__ == '__main__':
    create_example_excel()
